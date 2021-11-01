#########################################################
#対象はwatersからのexportテキストデータ
#########################################################

#重複のあるサンプル名に通し番号を振ることができない

import os,openpyxl, pyperclip
from openpyxl.styles.borders import Border, Side
import pandas as pd
import numpy as np

wb=openpyxl.Workbook("")
sheet = wb.create_sheet("HPLC")
num=3

#====================================
#基準となるrtを設定する関数
def input_rt():
    std_rt=input("基準となるrtを入力してください ")
    try:
        std_rt=float(std_rt)
        return std_rt
    except:
        print("エラー：数字を入力してください ")
        return input_rt()

#====================================
#テキストファイル内で戦闘が#の列数を返す
def return_sharp_nrows(txt):
    file=open(txt,"r",encoding="shift_jis")  #shift_jis形式だった
    text=file.read()
    text_split = text.split("\n")
    nrows = 0
    for l in text_split:
        if l.startswith("#"):
            sharp_nrow = nrows
        nrows += 1
    return sharp_nrow

#====================================
#フォルダ内からデータリストを作る関数
#data_dicのデータ構造は
#{ {file.text1: {"df1":df1(サンプル名,分析日,注入量,装置メソッドなど), "df2":df2(rt,area,area%など)}, file.text2:...}
def datalist_from_dirpath (folder):
    data_dic = {}
    os.chdir(folder)
    for file_txt in os.listdir():
        if not file_txt.endswith(".txt"):
            continue
        nrows = return_sharp_nrows(file_txt)
        df1=pd.read_csv(file_txt, header = None, nrows=nrows, encoding="shift_jis", sep="\t",  index_col=0)
        df2=pd.read_csv(file_txt, header = nrows, encoding="shift_jis", sep="\t", index_col="#")
        analysis_date = df1.loc["分析日"].values[0]
        data_dic[file_txt] = {"df1":df1, "df2":df2, "analysis_data":analysis_date}
    return data_dic

#====================================
#フォルダ内からデータを収集
folder=input("整理したいフォルダを入力してください")
data_dic=datalist_from_dirpath(folder)
#分析日でソートする
data_dic= {k: data_dic[k] for k in sorted(data_dic, reverse=False, key=lambda x: data_dic[x]["analysis_data"])}

#====================================
#rrtを設定し、rrtを収集する
rrt_list=pd.Series()
for data in data_dic:
    df1 = data_dic[data]["df1"]
    df2 = data_dic[data]["df2"]
    file=open(data,"r",encoding="shift_jis")  #dataは一番上のkey、つまりテキストファイル名
    text=file.read()                       #一度きれいなデータを見せて、rrtを決定させる
    print("================================================================")
    print(text)
    file.close()
    print("サンプル名：",df1.loc["サンプル名"].values[0])
    print("面積最大の保持時間：",df2["保持時間"][df2["面積"]==df2["面積"].max()].values[0], "min")
    print("クリップボードに貼り付けました。")
    pyperclip.copy(df2["保持時間"][df2["面積"]==df2["面積"].max()].values[0])
    std_rt = input_rt()
    print()
    df_rrt = df2["保持時間"]/std_rt
    df_rrt = df_rrt.round(2)
    df2["相対保持時間"] = df_rrt
    rrt_list=pd.concat([rrt_list, df2["相対保持時間"]])
rrt_list = np.sort(rrt_list.unique())


#====================================
#rrtをエクセルの左端に書き込む
for i,j in zip(rrt_list,range(4,len(rrt_list)+4)):
    sheet.cell(row=j,column=1).value=i

#========================================
#data_listをエクセルに書き込む
#罫線を引く
side = Side(style='thin')
border = Border(right=side)
bottom = Border(bottom=side)
for data in data_dic:
    df1 = data_dic[data]["df1"]
    df2 = data_dic[data]["df2"]
    sheet.cell(row=1,column=num).value="サンプル名"
    sheet.cell(row=2,column=num).value="分析日"
    sheet.cell(row=1,column=num+1).value=df1.loc["サンプル名"].values[0]
    sheet.cell(row=2,column=num+1).value=df1.loc["分析日"].values[0]
    sheet.cell(row=3,column=num).value="rrt"
    sheet.cell(row=3,column=num+1).value="rt"
    sheet.cell(row=3,column=num+2).value="area"
    sheet.cell(row=3,column=num+3).value="area%"
    sheet.cell(row=1,column=num+3).border = border
    sheet.cell(row=2,column=num+3).border = border
    sheet.cell(row=3,column=num+3).border = border
    for r in range(4,len(rrt_list)+4):
        sheet.cell(row=r,column=num+3).border = border
        for n in df2["相対保持時間"]:
            if n==sheet.cell(row=r,column=1).value:
                sheet.cell(row=r,column=num).value=n
                sheet.cell(row=r,column=num+1).value=df2[df2["相対保持時間"]==n]["保持時間"].values[0]
                sheet.cell(row=r,column=num+2).value=df2[df2["相対保持時間"]==n]["面積"].values[0]
                sheet.cell(row=r,column=num+3).value=df2[df2["相対保持時間"]==n]["％面積"].values[0]
            else:
                continue
    num+=4
sheet.freeze_panes="C4"
for i in range(1,num):
    sheet.cell(row=len(rrt_list)+3,column=i).border = bottom

#==========================================
#area%とareaをピックアップして別シートにまとめる
num1=3
sheet1 = wb.create_sheet("pick_up_area%")
sheet2 = wb.create_sheet("pick_up_area")

for i,j in zip(rrt_list,range(4,len(rrt_list)+4)):
    sheet1.cell(row=j,column=1).value=i
    sheet2.cell(row=j,column=1).value=i
sheet1.cell(row=j+1,column=1).value="sum of area%"
sheet2.cell(row=j+1,column=1).value="sum of area"
sheet1.cell(row=1,column=num1-1).value="サンプル名"
sheet1.cell(row=2,column=num1-1).value="分析日"
sheet2.cell(row=1,column=num1-1).value="サンプル名"
sheet2.cell(row=2,column=num1-1).value="分析日"
for data in data_dic:
    df1 = data_dic[data]["df1"]
    df2 = data_dic[data]["df2"]
    sheet1.cell(row=1,column=num1).value=df1.loc["サンプル名"].values[0]
    sheet1.cell(row=2,column=num1).value=df1.loc["分析日"].values[0]
    sheet1.cell(row=3,column=num1).value="area%"
    sheet2.cell(row=1,column=num1).value=df1.loc["サンプル名"].values[0]
    sheet2.cell(row=2,column=num1).value=df1.loc["分析日"].values[0]
    sheet2.cell(row=3,column=num1).value="area"
    for r in range(4,len(rrt_list)+4):
        for n in df2["相対保持時間"]:
            if n==sheet1.cell(row=r,column=1).value:
                sheet1.cell(row=r,column=num1).value=df2[df2["相対保持時間"]==n]["％面積"].values[0]
                sheet2.cell(row=r,column=num1).value=df2[df2["相対保持時間"]==n]["面積"].values[0]
            else:
                continue
    sheet1.cell(row=r+1,column=num1).value=df2["％面積"].sum()
    sheet2.cell(row=r+1,column=num1).value=df2["面積"].sum()
    num1+=1
sheet1.freeze_panes="C4"
sheet2.freeze_panes="C4"

#============================
#Sheetを削除、excelを保存、エクセルファイルの名前をフォルダから参照
wb.remove(wb["Sheet"])
wb.save(".\\{}.xlsx".format(os.path.splitext(os.path.basename(os.getcwd()))[0]))
wb.close()

#============================
#重複のあるファイルをリネーム
def rename (text_file, name):
    root, ext = os.path.splitext(text_file)
    if "/" in root:
        root = root.replace("/","_")
    try:
        num=0
        os.rename(text_file, name+ext)
    except FileNotFoundError:
        num +=1
        name = name + "{}".format(str(num))
        rename(text_file, name)

#============================
#ファイル名を変更する
import os.path
for data in data_dic:
    df1 = data_dic[data]["df1"]
    for d in os.listdir():
        if d[:-4]==data[:-4]:
            root, ext = os.path.splitext(d)
            sample_name = df1.loc["サンプル名"].values[0]
            if "/" in sample_name:
                sample_name = sample_name.replace("/","_")
            os.rename(d, sample_name + ext)

print()
print("================================")
print("エクセルファイルを作成しました。")
print("================================")