#########################################################
#日立の.xlsファイルからプロファイルを作成
#分析日ごとに並び替え
#pick_upも作成
#########################################################
print('ライブラリをロード中...\n')

import os,openpyxl,xlrd, pyperclip
from openpyxl.styles.borders import Border, Side
from glob import glob
import datetime
import pandas as pd
import numpy as np

def excel_date(num):   #シリアル値を日付に変換する
    from datetime import datetime, timedelta
    return(datetime(1899, 12, 30) + timedelta(days=num))

folder = input("整理したいフォルダのパスを入力してください")
os.chdir(folder)
file_list = glob(folder+'\\*.XLS')

data_list=[] # [{'analysistime':---, 'samplename':---, 'df_mr':pd.DataFrame(), 'df_rt':pd.DataFrame()}, {}, {}...]
rrt_list = pd.Series(dtype=float)

for file in file_list:
    df_mr = pd.read_excel(file, sheet_name='Manager Report')
    df_mr.columns = [i for i in range(df_mr.shape[1])]
    samplename = df_mr.iloc[8,5]
    analysistime = df_mr.iloc[3,2]
    
    df_tables = pd.read_excel(file, sheet_name='Tables')
    df_tables.columns = [i for i in range(df_mr.shape[1])]
    no_idx = df_tables.index[df_tables.iloc[:,1]=='NO'] # get index of NO in column 1.
    if len(no_idx)>=2:
        no_idx = no_idx[0]
        no_idx = pd.core.indexes.numeric.Int64Index(np.array([no_idx]))
    nan_idx = df_tables.index[df_tables.iloc[:,1].isna()] # get index of NaN in column 1.
    last_idx = nan_idx[nan_idx.values>no_idx.values]
    df_rt = df_tables.iloc[no_idx.values[0]+1:last_idx.values[0],1:]
    df_rt.columns = df_tables.iloc[no_idx.values[0],1:]
    
    max_idx = np.argmax(df_rt.iloc[:,3])
    max_rt = df_rt.iloc[max_idx, 1]
    pyperclip.copy(max_rt)
    
    print("================================================================")
    print("分析日:", analysistime)
    print("サンプル名:", samplename)
    print(df_rt)
    print('最大面積値のRT:', max_rt)
    print('最大面積値のRTをクリップボードに貼り付けました。')
    print()
    
    std_rt = float(input("基準となるrtを入力してください"))
    rrt = df_rt.iloc[:,1] / std_rt
    rrt = rrt.astype(float)
    rrt = rrt.round(2)
    df_rt['RRT'] = rrt
    rrt_list = pd.concat([rrt_list, rrt])
    
    data_list.append({'analysistime':analysistime, 'samplename':samplename, 'df_rt':df_rt})

data_list = sorted(data_list, key=lambda x: x['analysistime'])
rrt_list = np.sort(rrt_list.unique())

#========================================
# 不純物プロファイルを書くためのexcelファイルを用意する
print('\n不純物プロファイルのExcelファイルの作成中...')
wb1 = openpyxl.Workbook("")
sheet1 = wb1.create_sheet("HPLC")
num = 3

# rrtをエクセルの左端に書き込む
for i,j in zip(rrt_list,range(4,len(rrt_list)+4)):
    sheet1.cell(row=j,column=1).value=i   #左端にrrtを入力する 

side = Side(style='thin')
border = Border(right=side)
bottom = Border(bottom=side)
for m in data_list:#   mは[[サンプル名,分析日],[rrt,rt,area,area_percent],[rrt,rt,area,area_percent]...] 
    df_rt = m['df_rt']
    samplename = m['samplename']
    analysistime = m['analysistime']
    
    sheet1.cell(row=1,column=num).value="サンプル名"
    sheet1.cell(row=2,column=num).value="分析日"
    sheet1.cell(row=1,column=num+1).value=samplename #サンプル名代入
    sheet1.cell(row=2,column=num+1).value=analysistime  #分析日を代入
    sheet1.cell(row=3,column=num).value="rrt"
    sheet1.cell(row=3,column=num+1).value="rt"
    sheet1.cell(row=3,column=num+2).value="area"
    sheet1.cell(row=3,column=num+3).value="area%"
    sheet1.cell(row=1,column=num+3).border = border
    sheet1.cell(row=2,column=num+3).border = border
    sheet1.cell(row=3,column=num+3).border = border
    for r in range(4,len(rrt_list)+4):
        sheet1.cell(row=r,column=num+3).border = border
        for n in df_rt['RRT']:   # nは[サンプル名,分析日],[rrt,rt,area,area_percent],[rrt,rt,area,area_percent]...
            if n==sheet1.cell(row=r,column=1).value:
                sheet1.cell(row=r,column=num).value=n
                sheet1.cell(row=r,column=num+1).value=df_rt[df_rt['RRT']==n].iloc[:,1].values[0]
                sheet1.cell(row=r,column=num+2).value=df_rt[df_rt['RRT']==n].iloc[:,2].values[0]
                sheet1.cell(row=r,column=num+3).value=df_rt[df_rt['RRT']==n].iloc[:,3].values[0]
            else:
                continue
    num+=4
sheet1.freeze_panes="C4"
for i in range(1,num):
    sheet1.cell(row=len(rrt_list)+3,column=i).border = bottom
#==========================================

#==========================================
#area%をピックアップしてまとめる
num1=3
sheet1 = wb1.create_sheet("pick_up_area%")
sheet2 = wb1.create_sheet("pick_up_area")

for i,j in zip(rrt_list,range(4,len(rrt_list)+4)):
    sheet1.cell(row=j,column=1).value=i
    sheet2.cell(row=j,column=1).value=i
    sheet1.cell(row=j+1,column=1).value="sum of area%"
    sheet2.cell(row=j+1,column=1).value="sum of area"
    sheet1.cell(row=1,column=num1-1).value="サンプル名"
    sheet1.cell(row=2,column=num1-1).value="分析日"
    sheet2.cell(row=1,column=num1-1).value="サンプル名"
    sheet2.cell(row=2,column=num1-1).value="分析日"
for m in data_list:
    df_rt = m['df_rt']
    samplename = m['samplename']
    analysistime = m['analysistime']
    sheet1.cell(row=1,column=num1).value=samplename
    sheet1.cell(row=2,column=num1).value=analysistime
    sheet1.cell(row=3,column=num1).value="area%"
    sheet2.cell(row=1,column=num1).value=samplename
    sheet2.cell(row=2,column=num1).value=analysistime
    sheet2.cell(row=3,column=num1).value="area"
    for r in range(4,len(rrt_list)+4):
        for n in df_rt['RRT']:
            if n==sheet1.cell(row=r,column=1).value:
                sheet1.cell(row=r,column=num1).value=df_rt[df_rt['RRT']==n].iloc[:,3].values[0]
                sheet2.cell(row=r,column=num1).value=df_rt[df_rt['RRT']==n].iloc[:,2].values[0]
            else:
                continue
    sheet1.cell(row=r+1,column=num1).value=df_rt.iloc[:,2].sum()
    sheet2.cell(row=r+1,column=num1).value=df_rt.iloc[:,2].sum()
    num1+=1
sheet1.freeze_panes="C4"
sheet2.freeze_panes="C4"
#==========================================
dt_now = datetime.datetime.now()
dt_now = dt_now.strftime('%Y%m%d%H%M%S')
dirname = os.path.splitext(os.path.basename(os.getcwd()))[0]
filename = dirname + '_' + dt_now
wb1.remove(wb1["Sheet"])
wb1.save(".\\{}.xlsx".format(filename))
wb1.close()

print()
print("================================")
print("エクセルファイルを作成しました。")
print("================================")